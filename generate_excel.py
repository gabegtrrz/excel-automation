import logging
import pandas as pd
import numpy as np
from faker import Faker
import random
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

### Configuration
TICKER_SYMBOLS = ['AAPL', 'GOOGL', 'MSFT', 'AMZN', 'TSLA', 'META', 'NFLX', 'NVDA', 'INTC', 'AMD']

BASE_PRICES = {
    'AAPL' : 247.66,
    'GOOGL' : 245.45,
    'MSFT' : 513.57,
    'AMZN' : 216.39,
    'TSLA' : 429.24,
    'META' : 708.65,
    'NFLX' : 1215.35,
    'NVDA' : 180.03,
    'INTC' : 35.63,
    'AMD'  : 218.09
}

def configure_logging(level: int = logging.INFO) -> None:
    """Configure the root logger for the application.

    The default level is INFO; set to DEBUG for verbose diagnostics.
    """
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )

def validate_scenario_config(quantity_range, status_distribution) -> None:
    """Validate scenario inputs to prevent runtime errors and ensure sane defaults.

    Raises ValueError with descriptive messages when validation fails.
    """
    if not isinstance(quantity_range, (tuple, list)) or len(quantity_range) != 2:
        raise ValueError("quantity_range must be a (min, max) tuple/list with two elements")
    min_q, max_q = quantity_range
    if not (isinstance(min_q, (int, np.integer)) and isinstance(max_q, (int, np.integer))):
        raise ValueError("quantity_range values must be integers")
    if min_q < 0 or max_q <= 0 or min_q >= max_q:
        raise ValueError("quantity_range must be non-negative and min < max")

    if not isinstance(status_distribution, dict) or not status_distribution:
        raise ValueError("status_distribution must be a non-empty dict of {status: weight}")
    for status, weight in status_distribution.items():
        if not isinstance(status, str) or not status:
            raise ValueError("All status keys must be non-empty strings")
        if not isinstance(weight, (int, float)) or weight < 0:
            raise ValueError("All status weights must be non-negative numbers")
    if sum(status_distribution.values()) == 0:
        raise ValueError("Sum of status_distribution weights must be > 0")

def generate_synthetic_trade_data(num_records, quantity_range, status_distribution):
    """
    Generates a list of synthetic financial trade data records based on specific rules as a rule-based submodel.

    Args:
        num_records (int): Number of trade records to generate.
        quantity_range (tuple): Min and max quantity for trades.
        status_distribution (dict): dictionary defining the weighted distribution of statuses.

    - Uses Numpy for numerical data.
    - Uses Faker for realistic, non-numerical metadata (Source IP).
    """

    ### Initialization
    fake = Faker()
    trade_data = []
    statuses = list(status_distribution.keys())
    weights = list(status_distribution.values())


    for _ in range(num_records):
        # Use per-record try to avoid aborting the whole run due to one bad record
        try:
            ticker = random.choice(TICKER_SYMBOLS)
            base_price = BASE_PRICES[ticker]

            trade_price = round(np.random.normal(loc=base_price, scale=1.5), 2)
            quantity = np.random.randint(low=quantity_range[0], high=quantity_range[1])

            trade_status = random.choices(statuses, weights, k=1)[0] # always returns a list so take first element

            trade_data.append({
                'Trade ID': str(uuid.uuid4()),
                'Account ID': str(uuid.uuid4()),
                'Ticker': ticker,
                'Trade Type': random.choice(['BUY', 'SELL']),
                'Quantity': quantity,
                'Trade Price': trade_price,
                'Status': trade_status,
                'Source IP': fake.ipv4()
            })
        except Exception as e:
            logging.warning("Skipping record due to generation error: %s", e)
    return trade_data

def apply_excel_formatting(excel_file, num_records, large_trade_threshold=500000):
    """Applies professional formatting to the Excel file for an easy-to-read report."""

    # Load workbook and select active sheet
    try:
        wb = load_workbook(excel_file)
    except Exception as exc:
        logging.error("Failed to load workbook '%s': %s", excel_file, exc)
        return

    ws = wb.active

    if ws is None:
        logging.error("No active worksheet found in '%s'. Cannot apply formatting.", excel_file)
        return

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0a3e7d", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Styles for conditional formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for col_idx, all_col_cells in enumerate(ws.columns, 1):

        ### Format currency columns
        column_header = ws.cell(row=1, column=col_idx).value
        if column_header in ['Trade Price', 'Trade Value']:
            for cell in all_col_cells[1:]:
                cell.number_format = '$#,##0.00'

        ### Auto-adjust column width
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in all_col_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                # If a cell's value can't be stringified, skip width contribution
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ### Conditional Formatting Logic

    data_range = f"A2:{get_column_letter(ws.max_column)}{num_records + 1}"

    # Rule 1: highlight a row if the 'Status' in column H is "FAILED"
    failed_trade_rule = FormulaRule(formula=['$H2="FAILED"'], fill=red_fill, font=red_font)

    # Rule 2: highlight 'Trade Value' in column G if it exceeds a threshold (e.g., $500,000)
    large_trade_rule = FormulaRule(formula=[f'$G2>{large_trade_threshold}'], fill=yellow_fill)

    try:
        ws.conditional_formatting.add(data_range, failed_trade_rule)
        ws.conditional_formatting.add(data_range, large_trade_rule)
    except Exception as exc:
        logging.warning("Failed to add conditional formatting: %s", exc)

    try:
        wb.save(excel_file)
        logging.info("Applied formatting to '%s'.", excel_file)
    except Exception as exc:
        logging.error("Failed to save formatted workbook '%s': %s", excel_file, exc)

    return 




def main():
    """Main function to define rules, select a scenario, and generate the data.

    This 
    demonstrates building a configurable tool with robust error handling around generation and I/O steps.
    """

    configure_logging()
    logging.info("Starting Trade Data Excel File Generation...")

    ### Configuration
    NUM_RECORDS = 100
    OUTPUT_FILENAME = 'output.xlsx'
    LARGE_TRADE_THRESHOLD = 3000000 # for flagging large trades
    
    ### Define Rules for QA Scenario ###

    scenarios = {
        'standard day': {
            'quantity_range': (10, 5001),
            'status_distribution': {'EXECUTED': 94, 'PENDING': 4, 'FAILED': 2}
        },
        'high_volume_failures': {
            'quantity_range': (100, 10001),
            'status_distribution': {'EXECUTED': 60, 'PENDING': 10, 'FAILED': 30}
        },
        'institutional_trades': {
            'quantity_range': (50000, 200001), # testing large numbers
            'status_distribution': {'EXECUTED': 98, 'PENDING': 2, 'FAILED': 0}
        },
    }

    ### Select Scenario to run ###
    # QA tester would change this to get different data according to their needs
    selected_scenario_name = 'standard day'
    selected_scenario = scenarios[selected_scenario_name] # to double-check

    logging.info("Generating excel data for scenario: %s", selected_scenario)

    # Validate configuration before generating data to fail-fast on bad input
    try:
        validate_scenario_config(
            quantity_range=selected_scenario['quantity_range'],
            status_distribution=selected_scenario['status_distribution']
        )
    except ValueError as exc:
        logging.error("Invalid scenario configuration: %s", exc)
        return

    trade_data = generate_synthetic_trade_data(
        num_records=NUM_RECORDS,
        quantity_range=selected_scenario['quantity_range'],
        status_distribution=selected_scenario['status_distribution']
    )

    ### Process and save data to Excel
    df = pd.DataFrame(trade_data)

    ### FIRST LOGICAL OPERATION: DERIVED COLUMN
    df['Trade Value'] = df['Trade Price'] * df['Quantity']

    ### SECOND LOGICAL OPERATION: COMPLEX DERIVED COLUMN
    # Flag trades if they failed OR are unusually large. This simulates a compliance rule.
    df['Flagged for Review'] = (df['Status'] == 'FAILED') | (df['Trade Value'] > LARGE_TRADE_THRESHOLD)

    df = df[['Trade ID', 'Account ID', 'Ticker', 'Trade Type', 'Trade Price', 'Quantity','Trade Value', 'Status', 'Source IP', 'Flagged for Review']]

    try:
        df.to_excel(OUTPUT_FILENAME, index=False, sheet_name=selected_scenario_name)
        logging.info("Successfully generated '%s'.", OUTPUT_FILENAME)
    except Exception as exc:
        logging.error("Failed to write Excel file '%s': %s", OUTPUT_FILENAME, exc)
        return

    apply_excel_formatting(excel_file=OUTPUT_FILENAME,num_records=NUM_RECORDS, large_trade_threshold=LARGE_TRADE_THRESHOLD)
    logging.info("Script finished successfully!")



if __name__ == "__main__":
    try:
        main()
    except Exception:
        # Top-level guard to ensure any unexpected exception is logged with traceback
        logging.exception("Unhandled error during execution")
